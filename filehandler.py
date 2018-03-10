from pathlib import Path, PurePosixPath
from abc import ABCMeta, abstractmethod
from csv import DictReader as CSVDictReader
from openpyxl import Workbook, load_workbook


# James
class FileHandler:
    # James
    def __init__(self, file_name):
        self.filename = file_name
        self.file_type = None

    # James
    @staticmethod
    def get_file_name():
        cwd = './Saves/'
        for file in Path(cwd).iterdir():
            print(file)
        file = input("Which file do you wish to load? >>> ")
        filename = Path(cwd+file)
        return filename

    # James
    def file_exist(self):
        if self.filename.exists():
            return True
        else:
            return False

    # Wesley
    def set_file_type(self):
        """Will get the file type and will create the
            corresponding solid class and set it to self.file_type"""
        suffix = PurePosixPath('my/library/setup.py').suffix
        file_types = {
            '.csv': FileTypeCSV(),
            '.xlsx': FileTypeXLSX()
        }
        self.file_type = file_types[suffix]

    # James
    # def csv_read(self):
    #     with open(self.filename) as f:
    #         reader = csv.DictReader(f)
    #         for row in reader:
    #             print(row)


# Wesley
class FileTypeAbstract(metaclass=ABCMeta):
    # Wesley
    @abstractmethod
    def read(self, filename):
        pass


# Wesley
class FileTypeCSV(FileTypeAbstract):
    # James
    def read(self, filename):
        data = dict()
        empno = 0
        with open(self.filename) as f:
            reader = CSVDictReader(f)
            for row in reader:
                record = dict()
                for key in row:
                    record[key] = row.get(key)
                data[empno] = record
                empno += 1
            # print(data)
        return data


# Wesley
class FileTypeXLSX(FileTypeAbstract):
    # James
    def read(self, filename):
        data = dict()
        empno = 0
        keys = []
        workbook = load_workbook(filename)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        for row in worksheet.iter_rows():
            record = dict()
            for cell in row:
                if 'a' in cell:
                    keys.append(cell.value)
                else:
                    record[cell] = cell.value

        # with open(filename) as f:
        #     reader = CSVDictReader(f)
        #     for row in reader:
        #         print(row)



def run():
    a = FileHandler.load_file()
    aclass = FileHandler(a)

    while aclass.file_exist() is False:
        print("File exists:", aclass.file_exist())
        a = FileHandler.load_file()
        aclass = FileHandler(a)
    if aclass.file_exist() is True:
        aclass.csv_read()
